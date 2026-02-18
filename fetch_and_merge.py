#!/usr/bin/env python3
"""
Fetch data from Numberland API (getservice, getcountry, getinfo),
merge responses, and export to Excel.
"""

import json
import requests
from pathlib import Path
from datetime import datetime

# Optional: pandas + openpyxl for Excel. Install with: pip install requests pandas openpyxl
try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pd = None
    load_workbook = None

BASE_URL = "https://api.numberland.ir/v2.php/"
API_KEY = "7143e4c5a8173ca572232dcc15773cbc"


def fetch_json(url: str) -> list | dict:
    """GET URL and return parsed JSON."""
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    return resp.json()


def main():
    # 1) Fetch all three endpoints
    print("Fetching getservice...")
    services = fetch_json(
        f"{BASE_URL}?apikey={API_KEY}&method=getservice"
    )
    print("Fetching getcountry...")
    countries = fetch_json(
        f"{BASE_URL}?apikey={API_KEY}&method=getcountry"
    )
    print("Fetching getinfo...")
    info = fetch_json(
        f"{BASE_URL}?apikey={API_KEY}&method=getinfo&operator=&country="
    )

    # Normalize to lists if API returns wrapped object
    if isinstance(services, dict):
        services = services.get("data", services.get("result", [services]))
    if isinstance(countries, dict):
        countries = countries.get("data", countries.get("result", [countries]))
    if isinstance(info, dict):
        info = info.get("data", info.get("result", [info]))

    if not isinstance(services, list):
        services = [services] if services else []
    if not isinstance(countries, list):
        countries = [countries] if countries else []
    if not isinstance(info, list):
        info = [info] if info else []

    # 2) Lookup maps: id -> record
    service_by_id = {str(s.get("id")): s for s in services if s}
    country_by_id = {str(c.get("id")): c for c in countries if c}

    # 3) Merge: enrich each getinfo row with service and country details
    merged = []
    for row in info:
        if not isinstance(row, dict):
            continue
        sid = str(row.get("service", ""))
        cid = str(row.get("country", ""))
        svc = service_by_id.get(sid, {})
        cnt = country_by_id.get(cid, {})

        merged.append({
            **row,
            "service_name": svc.get("name", ""),
            "service_name_en": svc.get("name_en", ""),
            "country_name": cnt.get("name", row.get("cname", "")),
            "country_name_en": cnt.get("name_en", ""),
            "country_areacode": cnt.get("areacode", ""),
        })

    # 4) Add date column and export to Excel (filename = today's date)
    fetch_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    date_only = datetime.now().strftime("%Y-%m-%d")
    out_path = Path(__file__).parent / f"merged_numberland_{date_only}.xlsx"

    if pd is not None:
        df_new = pd.DataFrame(merged)
        df_new.insert(0, "date", fetch_date)

        if out_path.exists():
            df_old = pd.read_excel(out_path, engine="openpyxl", header=1)
            if "date" not in df_old.columns:
                df_old.insert(0, "date", "")
            df = pd.concat([df_old, df_new], ignore_index=True)
        else:
            df = df_new

        cols = ["date"] + [c for c in df.columns if c != "date"]
        df = df[cols]

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, startrow=1)
        wb = load_workbook(out_path)
        ws = wb.active
        ws.insert_rows(1)
        ws.cell(1, 1, f"Date: {fetch_date}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))

        # Styles: all text centered, bigger row heights
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Default row height for entire sheet (so all rows are taller)
        ws.sheet_format.defaultRowHeight = 30

        num_cols = len(cols)
        # Row 1: date header
        date_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for c in range(1, num_cols + 1):
            cell = ws.cell(1, c)
            cell.fill = date_fill
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[1].height = 42

        # Row 2: column headers
        header_fill = PatternFill(start_color="4A90A4", end_color="4A90A4", fill_type="solid")
        for c in range(1, num_cols + 1):
            cell = ws.cell(2, c)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[2].height = 36

        # Data rows: all columns centered, borders, zebra fill, taller rows
        light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for r in range(3, ws.max_row + 1):
            ws.row_dimensions[r].height = 30
            for c in range(1, num_cols + 1):
                cell = ws.cell(r, c)
                cell.border = thin_border
                cell.alignment = center_align  # center for every column
                cell.fill = light_fill if (r - 3) % 2 == 1 else white_fill

        # Column widths for readability
        for c in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 14

        wb.save(out_path)
        wb.close()
        print(f"Saved: {out_path} ({len(df)} total rows, {len(df_new)} new)")
    else:
        json_path = Path(__file__).parent / "merged_numberland.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(merged, f, ensure_ascii=False, indent=2)
        print(f"pandas/openpyxl not installed. Saved JSON: {json_path}")
        print("Install with: pip install pandas openpyxl")


if __name__ == "__main__":
    main()
